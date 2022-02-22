
import openpyxl
import string
import numpy as np
import pandas as pd

from numpy.linalg                     import norm
from string                           import digits
from unidecode                        import unidecode
from sklearn.feature_extraction.text  import CountVectorizer

def get_traceability(file_name):
    wb_obj = openpyxl.load_workbook(file_name) 
    sheet = wb_obj.active
    names_ls = [sheet['E' + str(i)].value[:-1] for i in range(2, sheet.max_row + 1)]
    billing_ls = [sheet['P' + str(i)].value for i in range(2, sheet.max_row + 1)]
    trace_dic = {'names': names_ls, 'billing': billing_ls}
    traceability = pd.DataFrame(trace_dic)
    return traceability

def cleanse(list_, diagnostic_folder='', order_folder=''):
    cleansed_list = []
    to_remove = [diagnostic_folder.lower(), order_folder.lower(), '_', '.pdf', '.',
                'orden y auto', 'autorizacion', 'caderas', 'rodillas',
                'rodilla', 'codo', 'orden', 'auto', 'hombro', 'c l s', 'cervical', 'pie',
                'columna', 'pie', 'carpograma', 'test', 'farril',
                'manos', 'torax', 'cadera', 'femur', 'antebrazo', 'brazo','  ']
    for row in list_:
        temp = row.lower()
        for remove_opt in to_remove: 
            temp = temp.replace(remove_opt, '')
            remove_digits = str.maketrans('', '', digits)
            temp = temp.translate(remove_digits)
            temp = unidecode(temp)
        cleansed_list.append(temp[1:])
    return cleansed_list

def get_Ngram(n):
  alph = list(string.ascii_lowercase)
  alph.extend(' ')
  n_gram = []
  if n == 1:
    n_gram.extend(alph)
  else: 
    for a in get_Ngram(n-1):
      for b in alph:
        n_gram.append(a+b)
  return n_gram

def get_voc_ngrams(n):
  vocab = [] 
  for i in range(n):
    j = i+1
    vocab.extend(get_Ngram(j))
  return vocab

def get_bag_grams(corpus, n):
    ngram_voctorizer  = CountVectorizer(ngram_range=(1,n), vocabulary = get_voc_ngrams(n), analyzer='char_wb')
    Xngram = ngram_voctorizer.fit_transform(corpus)
    Feat_ngram = ngram_voctorizer.get_feature_names_out()
    return Xngram, Feat_ngram

def get_ngram_matrix(cleansed_diagnostics, n):
    temp, _ = get_bag_grams(cleansed_diagnostics, n)
    return temp.todense()

def cosine_similarity(a, b): 
    a = np.ravel(a)
    b = np.ravel(b)
    return np.dot(a,b)/(norm(a)*norm(b))

def search_multiple_orders(cleansed_orders, orders_matrix, threshold):
    data_list = []
    elements_to_remove = []
    for i in range(len(cleansed_orders)):
        for j in range(i, len(cleansed_orders)):
            similarity = cosine_similarity(orders_matrix[i], orders_matrix[j])
            if similarity >= threshold and (i != j):
                temp = []
                # print("similarity: " + str(similarity) + " name_1: " + str(cleansed_orders[i])
                #                                         + " name_2: " + str(cleansed_orders[j]))
                temp.append(str(cleansed_orders[i]))
                temp.append(str(cleansed_orders[j]))
                temp.append(np.round(similarity, 3))
                data_list.append(temp)
                elements_to_remove.append(cleansed_orders[i])
                elements_to_remove.append(cleansed_orders[j])
    
    return data_list, elements_to_remove


def remove_multiple_orders(orders_remove, cleansed_orders, order_list_paths):
    cleansed_orders_temp = cleansed_orders.copy()
    order_list_paths_temp = order_list_paths.copy()
    indexes_order_remove = []
    removed_ls = []
    for order in orders_remove:
        if order in cleansed_orders_temp:
            indexes_order_remove.append(cleansed_orders_temp.index(order))
            removed_ls.append([order])
            cleansed_orders_temp.remove(order)

    for ix in indexes_order_remove:
        order_list_paths_temp.pop(ix)
    return cleansed_orders_temp, order_list_paths_temp, removed_ls

def get_pdf_order_pairs(cleansed_diagnostics, cleansed_orders, diagnostics_matrix, orders_matrix, threshold):
  indexes = []
  simil = []
  for i in range(len(cleansed_diagnostics)):
    for j in range(len(cleansed_orders)):
      similarity = cosine_similarity(orders_matrix[j], diagnostics_matrix[i])
      
      if similarity >= threshold:
        indexes.append([j, i])
        simil.append(similarity)

  min_index = np.argmin(simil)
  # print("Indexes length: "+ str(len(indexes)))S
  # print("Minimum similarity: "+ str(simil[min_index]))

  # print("Order: " + str(cleansed_orders[indexes[min_index][0]]))
  # print("PDF: " + str(cleansed_diagnostics[indexes[min_index][1]]))

  order_pd = []
  pdf_pd = []
  order_ix = []
  pdf_ix = []

  for ind in indexes:
    order_pd.append(cleansed_orders[ind[0]])
    pdf_pd.append(cleansed_diagnostics[ind[1]]) 
    order_ix.append(ind[0])
    pdf_ix.append(ind[1])

  data_c = pd.DataFrame({'Order_index': order_ix, 'Order': order_pd, 'PDF_index': pdf_ix, 'PDF': pdf_pd})
  metrics_ls = []

  metrics_ls.append(["Coincidencias detectadas", str(len(indexes))])
  metrics_ls.append(["Coincidencia menor", str(np.round(simil[min_index], 3))])
  metrics_ls.append(["Nombre en pdf", str(cleansed_diagnostics[indexes[min_index][1]])])
  metrics_ls.append(["Nombre en orden", str(cleansed_orders[indexes[min_index][0]])])
  metrics_ls.append(["Coincidencia global" , str(np.round(len(indexes)/len(cleansed_diagnostics),3))])

  matching = []

  for i in range(len(pdf_pd)):
    matching.append([pdf_pd[i], order_pd[i]])

  return data_c, metrics_ls, matching


def get_pdf_excel_pairs(traceability_pd, data_c, threshold):
  pdf_pd = data_c['PDF'].values
  pdf_pd_matrix = get_ngram_matrix(pdf_pd, 2)

  names = list(traceability_pd['names'].values)
  cleansed_names = cleanse(names)
  
  names_matrix = get_ngram_matrix(cleansed_names, 2)

  indexes = []
  simil = []
  i = 0
  j_list = []

  while i < len(pdf_pd_matrix):
    for j in range(i, len(names_matrix)):
      #print('i: ' + str(i) + ' j: ' + str(j))
      similarity = cosine_similarity(pdf_pd_matrix[i], names_matrix[j])
      if similarity >= threshold:
        #print('--X--')
        #print('i: ' + str(i) + ' j: ' + str(j) + '\n')
        
        #print('i: ' + str(i) + ' j: ' + str(j))
        if j not in j_list:
          # print('i: ' + str(i) + ' j: ' + str(j) + '\n')
          # print('similarity: ' + str(similarity) + ' pdf: ' + pdf_pd[i] + ' excel: ' + names[j] + '\n')
          indexes.append([i, j])
          simil.append(similarity)
          i += 1
          j_list.append(j)
          #print(i)
          # print('similarity: ' + str(similarity) + ' pdf: ' + pdf_pd[i] + ' excel: ' + names[j] + '\n')
          break

  # print("Indexes length: "+ str(len(indexes)))
  min_index = np.argmin(simil)
  # print("Minimum similarity: "+ str(simil[min_index]))

  # print("PDF: " + str(pdf_pd[indexes[min_index][0]]))
  # print("Names: " + str(cleansed_names[indexes[min_index][1]]))

  bill_n = traceability_pd['billing'].values
  metrics_ls = []

  metrics_ls.append(["Coincidencias detectadas", str(len(indexes))])
  metrics_ls.append(["Coincidencia menor", str(np.round(simil[min_index], 3))])
  metrics_ls.append(["Nombre en pdf", str(pdf_pd[indexes[min_index][0]])])
  metrics_ls.append(["Nombre en Excel", str(cleansed_names[indexes[min_index][1]])])
  metrics_ls.append(["Coincidencia global" , str(np.round(len(indexes)/len(pdf_pd),3))])
  
  matching = []

  for i in range(len(indexes)):
    matching.append([pdf_pd[indexes[i][0]], cleansed_names[indexes[i][1]]])

  return bill_n, metrics_ls, matching, indexes

def get_missing(indexes, traceability):
  names = list(traceability['names'].values)
  missing_ind = []
  for i in range(len(names)):
    if i not in np.array(indexes)[:,1]:
      missing_ind.append(i)
  missing_rows = traceability.iloc[missing_ind]
  
  missing_billings = []

  for i in range(len(missing_rows)):
    missing_billings.append([missing_rows['names'].iloc[i], missing_rows['billing'].iloc[i]])
  return missing_billings



